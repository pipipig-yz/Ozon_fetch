"""
Excel report writer for Ozon finance accruals.

Produces a multi-sheet workbook from pre-computed accrual data.
All styling and layout lives here; data transformation is delegated
to the ``transformations`` module.
"""

import sys

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from transformations import amt, extract_fee_rows, extract_product_rows

# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------

HEADER_FONT = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGN = Alignment(vertical="center")
AMOUNT_ALIGN = Alignment(horizontal="right", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
AMOUNT_FMT = "#,##0.00"

# ---------------------------------------------------------------------------
# Sheet headers
# ---------------------------------------------------------------------------

# Sheet "按操作排列" — Focused Summary
S0_HEADERS = [
    "应计ID", "日期", "单号", "SKU",
    "收入", "收单", "销售佣金", "配送佣金", "国际配送",
]

# Sheet "应计记录" — Accruals Summary
S1_HEADERS = [
    "应计ID", "日期", "发货单号", "配送模式", "应计类别",
    "总金额", "币种", "非商品费用ID", "非商品费用",
]

# Sheet "费用明细" — Fees Breakdown
S2_HEADERS = [
    "应计ID", "日期", "发货单号", "SKU",
    "费用类型ID", "费用类型名称", "金额", "币种",
    "来源",  # "item_fees" | "non_item_fee" | "delivery_service"
]

# Sheet "商品佣金" — Products Commission
S3_HEADERS = [
    "应计ID", "日期", "发货单号", "SKU",
    "销售金额", "售价(seller_price)",
    "销售佣金(sale_commission)", "平台佣金(commission)",
    "佣金比率", "奖金(bonus)", "共同投资(coinvestment)",
    "配送总费用", "币种",
]

# Sheet "费用类型" — Accrual Types
S4_HEADERS = ["类型ID", "名称", "描述"]

# Sheet "按订单号排列" — By Order
S_ORDER_HEADERS = [
    "基础单号", "日期", "SKU",
    "收入", "收单", "销售佣金", "配送佣金", "国际配送",
    "关联单号",
]

# ---------------------------------------------------------------------------
# Cell helpers
# ---------------------------------------------------------------------------


def _style_header(ws, row: int, num_cols: int) -> None:
    """Apply the standard header style to *num_cols* cells in *row*."""
    for col in range(1, num_cols + 1):
        c = ws.cell(row=row, column=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN
        c.border = THIN_BORDER


def _write_data_cell(
    ws, row: int, col: int, value, *, is_amount: bool = False
) -> None:
    """Write a data cell with standard body styling."""
    c = ws.cell(row=row, column=col, value=value)
    c.alignment = AMOUNT_ALIGN if is_amount else CELL_ALIGN
    c.border = THIN_BORDER
    if is_amount and isinstance(value, (int, float)):
        c.number_format = AMOUNT_FMT


def _auto_width(ws, widths: dict[int, float]) -> None:
    """Set explicit column widths from a ``{col_index: width}`` map."""
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------


def write_excel(
    accruals: list[dict],
    types_map: dict[int, dict],
    all_focused_rows: list[dict],
    order_groups: dict[str, dict],
    filepath: str,
) -> None:
    """Create and save the multi-sheet Excel workbook.

    Args:
        accruals: Raw accrual records from the API.
        types_map: ``{type_id: {name, description}}`` mapping.
        all_focused_rows: Pre-computed focused rows (one per accrual×SKU).
        order_groups: Pre-computed ``{base_pn: {field→value, dates, skus, pns}}``.
        filepath: Destination ``.xlsx`` path.
    """
    wb = Workbook()

    # ===================================================================
    # Sheet 1 — 按订单号排列
    # ===================================================================
    ws_order = wb.active
    ws_order.title = "按订单号排列"
    for col_idx, h in enumerate(S_ORDER_HEADERS, 1):
        ws_order.cell(row=1, column=col_idx, value=h)
    _style_header(ws_order, 1, len(S_ORDER_HEADERS))

    row = 2
    sorted_orders = sorted(
        order_groups.items(),
        key=lambda x: (max(x[1]["dates"]) if x[1]["dates"] else "", x[0]),
    )
    for base, g in sorted_orders:
        sorted_dates = sorted(g["dates"])
        if len(sorted_dates) == 1:
            date_str = sorted_dates[0]
        else:
            date_str = f"{sorted_dates[0]}~{sorted_dates[-1]}"
        sku_str = ", ".join(sorted(g["skus"])) if g["skus"] else "-"
        vals = [
            base,
            date_str,
            sku_str,
            g["收入"], g["收单"], g["销售佣金"],
            g["配送佣金"], g["国际配送"],
            ", ".join(sorted(g["pns"])),
        ]
        for col_idx, v in enumerate(vals, 1):
            _write_data_cell(ws_order, row, col_idx, v, is_amount=(4 <= col_idx <= 8))
        row += 1

    _auto_width(ws_order, {
        1: 24, 2: 22, 3: 22, 4: 14, 5: 14, 6: 14, 7: 14, 8: 14, 9: 40,
    })
    ws_order.freeze_panes = "A2"
    ws_order.auto_filter.ref = ws_order.dimensions

    # ===================================================================
    # Sheet 2 — 按操作排列
    # ===================================================================
    ws_summary = wb.create_sheet("按操作排列")
    for col_idx, h in enumerate(S0_HEADERS, 1):
        ws_summary.cell(row=1, column=col_idx, value=h)
    _style_header(ws_summary, 1, len(S0_HEADERS))

    for i, fr in enumerate(all_focused_rows, 2):
        vals = [
            fr["accrual_id"], fr["date"], fr["posting_number"],
            fr["sku"],
            fr["收入"], fr["收单"], fr["销售佣金"], fr["配送佣金"], fr["国际配送"],
        ]
        for col_idx, v in enumerate(vals, 1):
            _write_data_cell(ws_summary, i, col_idx, v, is_amount=(col_idx >= 5))

    _auto_width(ws_summary, {
        1: 14, 2: 12, 3: 22, 4: 12, 5: 16, 6: 16, 7: 16, 8: 16, 9: 16,
    })
    ws_summary.freeze_panes = "A2"
    ws_summary.auto_filter.ref = ws_summary.dimensions

    # ===================================================================
    # Sheet 3 — 应计记录
    # ===================================================================
    ws1 = wb.create_sheet("应计记录")
    for col_idx, h in enumerate(S1_HEADERS, 1):
        ws1.cell(row=1, column=col_idx, value=h)
    _style_header(ws1, 1, len(S1_HEADERS))

    row = 2
    for a in accruals:
        posting = a.get("posting") or {}
        nif = a.get("non_item_fee") or {}
        ta = a.get("total_amount") or {}
        vals = [
            a.get("accrual_id", ""),
            a.get("date", ""),
            a.get("unit_number", ""),
            posting.get("delivery_schema", ""),
            a.get("accrued_category", ""),
            amt(ta),
            ta.get("currency", ""),
            nif.get("type_id", ""),
            amt(nif.get("accrued")),
        ]
        for col_idx, v in enumerate(vals, 1):
            _write_data_cell(ws1, row, col_idx, v, is_amount=(col_idx in (6, 9)))
        row += 1

    _auto_width(ws1, {
        1: 14, 2: 12, 3: 22, 4: 14, 5: 18, 6: 16, 7: 8, 8: 14, 9: 16,
    })
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = ws1.dimensions

    # ===================================================================
    # Sheet 4 — 费用明细
    # ===================================================================
    ws2 = wb.create_sheet("费用明细")
    for col_idx, h in enumerate(S2_HEADERS, 1):
        ws2.cell(row=1, column=col_idx, value=h)
    _style_header(ws2, 1, len(S2_HEADERS))

    row = 2
    for a in accruals:
        for fr in extract_fee_rows(a, types_map):
            vals = [
                fr["accrual_id"], fr["date"], fr["posting_number"],
                fr["sku"], fr["type_id"], fr["type_name"],
                fr["amount"], fr["currency"], fr["source"],
            ]
            for col_idx, v in enumerate(vals, 1):
                _write_data_cell(ws2, row, col_idx, v, is_amount=(col_idx == 7))
            row += 1

    _auto_width(ws2, {
        1: 14, 2: 12, 3: 22, 4: 12, 5: 12, 6: 28, 7: 16, 8: 8, 9: 16,
    })
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions

    # ===================================================================
    # Sheet 5 — 商品佣金
    # ===================================================================
    ws3 = wb.create_sheet("商品佣金")
    for col_idx, h in enumerate(S3_HEADERS, 1):
        ws3.cell(row=1, column=col_idx, value=h)
    _style_header(ws3, 1, len(S3_HEADERS))

    row = 2
    for a in accruals:
        for pr in extract_product_rows(a):
            vals = [
                pr["accrual_id"], pr["date"], pr["posting_number"],
                pr["sku"],
                pr["sale_amount"], pr["seller_price"],
                pr["sale_commission"], pr["commission"],
                pr["commission_ratio"], pr["bonus"],
                pr["coinvestment"], pr["delivery_total"],
                pr["currency"],
            ]
            for col_idx, v in enumerate(vals, 1):
                is_amt = col_idx in (5, 6, 7, 8, 10, 11, 12)
                _write_data_cell(ws3, row, col_idx, v, is_amount=is_amt)
            row += 1

    _auto_width(ws3, {
        1: 14, 2: 12, 3: 22, 4: 12, 5: 16, 6: 16, 7: 16,
        8: 14, 9: 10, 10: 14, 11: 16, 12: 14, 13: 8,
    })
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = ws3.dimensions

    # ===================================================================
    # Sheet 6 — 费用类型
    # ===================================================================
    ws4 = wb.create_sheet("费用类型")
    for col_idx, h in enumerate(S4_HEADERS, 1):
        ws4.cell(row=1, column=col_idx, value=h)
    _style_header(ws4, 1, len(S4_HEADERS))

    row = 2
    for tid in sorted(types_map.keys()):
        t = types_map[tid]
        _write_data_cell(ws4, row, 1, tid)
        _write_data_cell(ws4, row, 2, t["name"])
        _write_data_cell(ws4, row, 3, t["description"])
        row += 1

    _auto_width(ws4, {1: 10, 2: 35, 3: 60})
    ws4.freeze_panes = "A2"

    # ---- Save ----
    wb.save(filepath)
    print(f"Saved Excel report to {filepath}", file=sys.stderr)

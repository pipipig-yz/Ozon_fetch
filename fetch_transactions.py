"""
Fetch Ozon finance accruals via v1 LTS endpoints.

Endpoints used (all v1, long-term stable):
    POST /v1/finance/accrual/types     —  type_id → name mapping
    POST /v1/finance/accrual/by-day    —  accrued charges per day (cursor pagination)

API docs: https://docs.ozon.ru/api/seller/zh/

Credentials are read from environment variables:
    OZON_CLIENT_ID  — your Ozon seller Client-Id
    OZON_API_KEY    — your Ozon seller Api-Key

Usage:
    python fetch_transactions.py                           # last 7 days → report.xlsx
    python fetch_transactions.py --days 30                 # last 30 days
    python fetch_transactions.py --from 2026-05-01 --to 2026-07-05
    python fetch_transactions.py --output report.xlsx      # explicit output path
    python fetch_transactions.py --json result.json        # also save raw JSON
"""

import argparse
import json
import os
import sys
from datetime import datetime, timedelta, timezone
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

BASE_URL = "https://api-seller.ozon.ru"
TYPES_ENDPOINT = "/v1/finance/accrual/types"
BY_DAY_ENDPOINT = "/v1/finance/accrual/by-day"
TIMEOUT = 30


# ===========================================================================
# Authentication
# ===========================================================================


def get_client() -> tuple[str, str]:
    """Read Client-Id and Api-Key from environment or a local .env file."""
    client_id = os.getenv("OZON_CLIENT_ID")
    api_key = os.getenv("OZON_API_KEY")

    if not client_id or not api_key:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        env_path = os.path.join(script_dir, ".env")
        if os.path.isfile(env_path):
            with open(env_path) as f:
                for line in f:
                    line = line.strip()
                    if line.startswith("#") or "=" not in line:
                        continue
                    k, v = line.split("=", 1)
                    k, v = k.strip(), v.strip().strip('"').strip("'")
                    if k == "OZON_CLIENT_ID" and not client_id:
                        client_id = v
                    elif k == "OZON_API_KEY" and not api_key:
                        api_key = v

    if not client_id or not api_key:
        sys.exit(
            "Missing credentials. Set OZON_CLIENT_ID and OZON_API_KEY env vars,\n"
            "or create a .env file next to this script with:\n"
            "    OZON_CLIENT_ID=your_client_id\n"
            "    OZON_API_KEY=your_api_key\n"
        )
    return client_id, api_key


# ===========================================================================
# API calls
# ===========================================================================


def _post(client_id: str, api_key: str, endpoint: str, body: dict) -> dict:
    """Generic POST to an Ozon API endpoint."""
    url = f"{BASE_URL}{endpoint}"
    headers = {
        "Client-Id": client_id,
        "Api-Key": api_key,
        "Content-Type": "application/json",
    }
    resp = requests.post(url, headers=headers, json=body, timeout=TIMEOUT)
    resp.raise_for_status()
    return resp.json()


def fetch_accrual_types(client_id: str, api_key: str) -> dict[int, dict]:
    """Return {type_id: {name, description}} mapping."""
    data = _post(client_id, api_key, TYPES_ENDPOINT, {})
    types_list = data.get("accrual_types") or (data.get("result") or {}).get("accrual_types", [])
    mapping: dict[int, dict] = {}
    for t in types_list:
        tid = t.get("id")
        if tid is not None:
            mapping[tid] = {"name": t.get("name", ""), "description": t.get("description", "")}
    return mapping


def fetch_accruals_for_date(
    client_id: str,
    api_key: str,
    date_str: str,
) -> list[dict]:
    """Fetch all accruals for a single date, following the last_id cursor."""
    all_accruals: list[dict] = []
    last_id: str = ""

    while True:
        body: dict = {"date": date_str}
        if last_id:
            body["last_id"] = last_id

        data = _post(client_id, api_key, BY_DAY_ENDPOINT, body)
        result = data.get("result", data)
        accruals = result.get("accruals", [])
        if not accruals:
            break

        all_accruals.extend(accruals)
        last_id = result.get("last_id", "")
        if not last_id:
            break

    return all_accruals


def fetch_all_accruals(
    client_id: str,
    api_key: str,
    date_from: str,
    date_to: str,
) -> list[dict]:
    """Fetch accruals for a date range, one day at a time."""
    all_accruals: list[dict] = []

    start = datetime.fromisoformat(date_from).date()
    end = datetime.fromisoformat(date_to).date()
    total_days = (end - start).days + 1

    for i in range(total_days):
        day = (start + timedelta(days=i)).isoformat()
        print(f"Fetching {day} ({i+1}/{total_days}) …", file=sys.stderr)
        day_accruals = fetch_accruals_for_date(client_id, api_key, day)
        all_accruals.extend(day_accruals)
        print(f"  → {len(day_accruals)} accruals", file=sys.stderr)

    return all_accruals


# ===========================================================================
# Excel writer
# ===========================================================================

HEADER_FONT = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGN = Alignment(vertical="center")
AMOUNT_ALIGN = Alignment(horizontal="right", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
AMOUNT_FMT = "#,##0.00"
SECTION_FONT = Font(name="Microsoft YaHei", bold=True, size=12, color="1F4E79")
LABEL_FONT = Font(name="Microsoft YaHei", bold=True)

# ---- Sheet 1: Accruals Summary ----
S1_HEADERS = [
    "应计ID",          # accrual_id
    "日期",            # date
    "发货单号",        # posting_number (unit_number)
    "配送模式",        # delivery_schema
    "应计类别",        # accrued_category
    "总金额",          # total_amount.amount
    "币种",            # total_amount.currency
    "非商品费用ID",    # non_item_fee.type_id
    "非商品费用",      # non_item_fee.accrued.amount
]

# ---- Sheet 2: Fees Breakdown (item_fees + non_item_fee expanded) ----
S2_HEADERS = [
    "应计ID",
    "日期",
    "发货单号",
    "SKU",
    "费用类型ID",
    "费用类型名称",
    "金额",
    "币种",
    "来源",      # "item_fees" | "non_item_fee" | "delivery_service"
]

# ---- Sheet 3: Products Commission ----
S3_HEADERS = [
    "应计ID",
    "日期",
    "发货单号",
    "SKU",
    "销售金额",               # sale_amount (收入)
    "售价(seller_price)",
    "销售佣金(sale_commission)",
    "平台佣金(commission)",
    "佣金比率",
    "奖金(bonus)",
    "共同投资(coinvestment)",
    "配送总费用",
    "币种",
]

# ---- Sheet 4: Accrual Types ----
S4_HEADERS = ["类型ID", "名称", "描述"]


def _style_header(ws, row: int, num_cols: int) -> None:
    for col in range(1, num_cols + 1):
        c = ws.cell(row=row, column=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN
        c.border = THIN_BORDER


def _write_data_cell(ws, row: int, col: int, value, *, is_amount: bool = False) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.alignment = AMOUNT_ALIGN if is_amount else CELL_ALIGN
    c.border = THIN_BORDER
    if is_amount and isinstance(value, (int, float)):
        c.number_format = AMOUNT_FMT


def _amt(d: dict, default=0.0) -> float:
    """Extract numeric amount from an Ozon money object {amount, currency}."""
    if not isinstance(d, dict):
        return default
    try:
        return float(d.get("amount", default))
    except (ValueError, TypeError):
        return default


def _auto_width(ws, widths: dict[int, float]) -> None:
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w


# ---- helpers to extract nested data ----


def _extract_fee_rows(accrual: dict, types_map: dict[int, dict]) -> list[dict]:
    """Expand item_fees + non_item_fee + delivery.services into flat rows."""
    rows: list[dict] = []
    base = {
        "accrual_id": accrual.get("accrual_id", ""),
        "date": accrual.get("date", ""),
        "posting_number": accrual.get("unit_number", ""),
    }

    # 1) item_fees → per-SKU fees (收单, 配送佣金, etc.)
    item_fees = accrual.get("item_fees") or {}
    for sku_group in (item_fees.get("fees") or []):
        sku = sku_group.get("sku", "")
        for fee in (sku_group.get("fees") or []):
            tid = fee.get("type_id", "")
            rows.append({
                **base,
                "sku": sku,
                "type_id": tid,
                "type_name": types_map.get(tid, {}).get("name", f"type_{tid}"),
                "amount": _amt(fee.get("accrued")),
                "currency": (fee.get("accrued") or {}).get("currency", ""),
                "source": "item_fees",
            })

    # 2) non_item_fee
    nif = accrual.get("non_item_fee") or {}
    if nif:
        tid = nif.get("type_id", "")
        rows.append({
            **base,
            "sku": "-",
            "type_id": tid,
            "type_name": types_map.get(tid, {}).get("name", f"type_{tid}"),
            "amount": _amt(nif.get("accrued")),
            "currency": (nif.get("accrued") or {}).get("currency", ""),
            "source": "non_item_fee",
        })

    # 3) delivery.services per product (国际配送 etc.)
    posting = accrual.get("posting") or {}
    for prod in (posting.get("products") or []):
        psku = prod.get("sku", "")
        delivery = prod.get("delivery") or {}
        for svc in (delivery.get("services") or []):
            tid = svc.get("type_id", "")
            rows.append({
                **base,
                "sku": psku,
                "type_id": tid,
                "type_name": types_map.get(tid, {}).get("name", f"type_{tid}"),
                "amount": _amt(svc.get("accrued")),
                "currency": (svc.get("accrued") or {}).get("currency", ""),
                "source": "delivery_service",
            })

    return rows


def _extract_product_rows(accrual: dict) -> list[dict]:
    """Flatten one accrual's posting.products into a list of dict rows."""
    rows: list[dict] = []
    base = {
        "accrual_id": accrual.get("accrual_id", ""),
        "date": accrual.get("date", ""),
        "posting_number": accrual.get("unit_number", ""),
    }
    posting = accrual.get("posting") or {}
    products = posting.get("products") or []
    for prod in products:
        sku = prod.get("sku", "")
        commission = prod.get("commission") or {}
        delivery = prod.get("delivery") or {}
        rows.append({
            **base,
            "sku": sku,
            "seller_price": _amt(commission.get("seller_price")),
            "sale_amount": _amt(commission.get("sale_amount")),
            "commission": _amt(commission.get("commission")),
            "commission_ratio": commission.get("commission_ratio", ""),
            "sale_commission": _amt(commission.get("sale_commission")),
            "bonus": _amt(commission.get("bonus")),
            "coinvestment": _amt(commission.get("coinvestment")),
            "delivery_total": _amt(delivery.get("total_accrued")),
            "currency": (commission.get("sale_amount") or {}).get("currency", ""),
        })
    return rows


def _match_fee(accrual: dict, sku: int, types_map: dict[int, dict], keywords: list[str],
               include_non_item: bool = True) -> float:
    """Sum fees in item_fees + non_item_fee + delivery.services whose type name matches."""
    total = 0.0
    kw_lower = [k.lower() for k in keywords]

    def _hit(tid) -> bool:
        return any(kw in types_map.get(tid, {}).get("name", "").lower() for kw in kw_lower)

    # item_fees (per-SKU)
    for sg in ((accrual.get("item_fees") or {}).get("fees") or []):
        if (sg.get("sku") or 0) != sku:
            continue
        for fee in (sg.get("fees") or []):
            if _hit(fee.get("type_id", 0)):
                total += _amt(fee.get("accrued"))

    # non_item_fee (not per-SKU — only count once across all SKU rows)
    if include_non_item:
        nif = accrual.get("non_item_fee") or {}
        if nif and _hit(nif.get("type_id", 0)):
            total += _amt(nif.get("accrued"))

    # delivery.services (per-SKU)
    for prod in ((accrual.get("posting") or {}).get("products") or []):
        if (prod.get("sku") or 0) != sku:
            continue
        for svc in ((prod.get("delivery") or {}).get("services") or []):
            if _hit(svc.get("type_id", 0)):
                total += _amt(svc.get("accrued"))
    return total


def _get_commission_val(accrual: dict, sku: int, field: str) -> float:
    """Get a commission field value for a specific SKU."""
    for prod in ((accrual.get("posting") or {}).get("products") or []):
        if prod.get("sku") == sku:
            return _amt((prod.get("commission") or {}).get(field))
    return 0.0


def _build_focused_rows(accrual: dict, types_map: dict[int, dict]) -> list[dict]:
    """One row per (accrual, SKU) with 收入/收单/销售佣金/配送佣金/国际配送."""
    rows: list[dict] = []
    base = {
        "accrual_id": accrual.get("accrual_id", ""),
        "date": accrual.get("date", ""),
        "posting_number": accrual.get("unit_number", ""),
    }
    # Collect all SKUs from products + item_fees
    # Use `or 0` (not `.get("sku", 0)`) because JSON null → Python None,
    # and None would break sorted(set) below.
    skus: set[int] = set()
    for prod in ((accrual.get("posting") or {}).get("products") or []):
        skus.add(prod.get("sku") or 0)
    for sg in ((accrual.get("item_fees") or {}).get("fees") or []):
        skus.add(sg.get("sku") or 0)
    skus.discard(0)

    if not skus:
        # No SKU-level data: one row with just non-item info
        rows.append({
            **base,
            "sku": "-",
            "收入": 0.0,
            "收单": _match_fee(accrual, 0, types_map, ["acquiring"]),
            "销售佣金": 0.0,
            "配送佣金": _match_fee(accrual, 0, types_map, ["lastmile", "logistic", "rfbsdomesticdelivery", "rfbsdomesticagentfee", "rfbsglobalagentfee", "shipment", "fulfillment"]),
            "国际配送": _match_fee(accrual, 0, types_map, ["rfbsglobaldelivery", "internationallogistic", "ozongloballogistics"]),
        })
    else:
        sorted_skus = sorted(skus)
        for i, sku in enumerate(sorted_skus):
            inc_nif = (i == 0)  # non_item_fee only on first SKU row
            rows.append({
                **base,
                "sku": sku,
                "收入": _get_commission_val(accrual, sku, "sale_amount"),
                "收单": _match_fee(accrual, sku, types_map, ["acquiring"], include_non_item=inc_nif),
                "销售佣金": _get_commission_val(accrual, sku, "sale_commission"),
                "配送佣金": _match_fee(accrual, sku, types_map, ["lastmile", "logistic", "rfbsdomesticdelivery", "rfbsdomesticagentfee", "rfbsglobalagentfee", "shipment", "fulfillment"], include_non_item=inc_nif),
                "国际配送": _match_fee(accrual, sku, types_map, ["rfbsglobaldelivery", "internationallogistic", "ozongloballogistics"], include_non_item=inc_nif),
            })
    return rows


# ---- Sheet 0: Focused Summary headers ----
S0_HEADERS = [
    "应计ID",
    "日期",
    "单号",
    "SKU",
    "收入",
    "收单",
    "销售佣金",
    "配送佣金",
    "国际配送",
]


def write_excel(
    accruals: list[dict],
    types_map: dict[int, dict],
    filepath: str,
) -> None:
    wb = Workbook()

    # ---- Pre-compute focused rows (needed by both Orders and Summary) ----
    all_focused_rows: list[dict] = []
    for a in accruals:
        for fr in _build_focused_rows(a, types_map):
            all_focused_rows.append(fr)

    # ===================================================================
    # Sheet 1 — 按订单号排列 (每个基础单号一行)
    # ===================================================================
    import re
    S_ORDER_HEADERS = [
        "基础单号",
        "日期",
        "SKU",
        "收入",
        "收单",
        "销售佣金",
        "配送佣金",
        "国际配送",
        "关联单号",
    ]

    def _base_posting_number(pn: str) -> str:
        return re.sub(r'-\d{1,2}$', '', str(pn))

    order_groups: dict[str, dict] = {}
    for fr in all_focused_rows:
        pn = str(fr["posting_number"])
        base = _base_posting_number(pn)
        if base not in order_groups:
            order_groups[base] = {
                "收入": 0.0, "收单": 0.0, "销售佣金": 0.0,
                "配送佣金": 0.0, "国际配送": 0.0,
                "dates": set(), "skus": set(), "pns": set(),
            }
        g = order_groups[base]
        for field in ["收入", "收单", "销售佣金", "配送佣金", "国际配送"]:
            g[field] += fr.get(field, 0.0) or 0.0
        if fr.get("date"):
            g["dates"].add(str(fr["date"]))
        sku = str(fr.get("sku", "-"))
        if sku and sku != "-":
            g["skus"].add(sku)
        g["pns"].add(pn)

    ws_order = wb.active
    ws_order.title = "按订单号排列"
    for col_idx, h in enumerate(S_ORDER_HEADERS, 1):
        ws_order.cell(row=1, column=col_idx, value=h)
    _style_header(ws_order, 1, len(S_ORDER_HEADERS))

    row = 2
    for base in sorted(order_groups.keys()):
        g = order_groups[base]
        sorted_dates = sorted(g["dates"])
        if len(sorted_dates) == 1:
            date_str = sorted_dates[0]
        else:
            date_str = f"{sorted_dates[0]}~{sorted_dates[-1]}"
        sku_str = ", ".join(sorted(g["skus"])) if g["skus"] else "-"
        vals = [
            base,
            date_str,
            sku_str,
            g["收入"], g["收单"], g["销售佣金"],
            g["配送佣金"], g["国际配送"],
            ", ".join(sorted(g["pns"])),
        ]
        for col_idx, v in enumerate(vals, 1):
            _write_data_cell(ws_order, row, col_idx, v, is_amount=(4 <= col_idx <= 8))
        row += 1

    _auto_width(ws_order, {
        1: 24, 2: 22, 3: 22, 4: 14, 5: 14, 6: 14, 7: 14, 8: 14, 9: 40,
    })
    ws_order.freeze_panes = "A2"
    ws_order.auto_filter.ref = ws_order.dimensions

    # ===================================================================
    # Sheet 2 — 按操作排列 (每条应计记录展开为一行)
    # ===================================================================
    ws_summary = wb.create_sheet("按操作排列")
    for col_idx, h in enumerate(S0_HEADERS, 1):
        ws_summary.cell(row=1, column=col_idx, value=h)
    _style_header(ws_summary, 1, len(S0_HEADERS))

    for i, fr in enumerate(all_focused_rows, 2):
        vals = [
            fr["accrual_id"], fr["date"], fr["posting_number"],
            fr["sku"],
            fr["收入"], fr["收单"], fr["销售佣金"], fr["配送佣金"], fr["国际配送"],
        ]
        for col_idx, v in enumerate(vals, 1):
            _write_data_cell(ws_summary, i, col_idx, v, is_amount=(col_idx >= 5))

    _auto_width(ws_summary, {
        1: 14, 2: 12, 3: 22, 4: 12, 5: 16, 6: 16, 7: 16, 8: 16, 9: 16,
    })
    ws_summary.freeze_panes = "A2"
    ws_summary.auto_filter.ref = ws_summary.dimensions

    # ===================================================================
    # Sheet 3 — 应计记录
    # ===================================================================
    ws1 = wb.create_sheet("应计记录")
    for col_idx, h in enumerate(S1_HEADERS, 1):
        ws1.cell(row=1, column=col_idx, value=h)
    _style_header(ws1, 1, len(S1_HEADERS))

    row = 2
    for a in accruals:
        posting = a.get("posting") or {}
        nif = a.get("non_item_fee") or {}
        ta = a.get("total_amount") or {}

        vals = [
            a.get("accrual_id", ""),
            a.get("date", ""),
            a.get("unit_number", ""),
            posting.get("delivery_schema", ""),
            a.get("accrued_category", ""),
            _amt(ta),
            ta.get("currency", ""),
            nif.get("type_id", ""),
            _amt(nif.get("accrued")),
        ]
        for col_idx, v in enumerate(vals, 1):
            is_amt = col_idx in (6, 9)
            _write_data_cell(ws1, row, col_idx, v, is_amount=is_amt)
        row += 1

    _auto_width(ws1, {
        1: 14, 2: 12, 3: 22, 4: 14, 5: 18, 6: 16, 7: 8, 8: 14, 9: 16,
    })
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = ws1.dimensions

    # ===================================================================
    # Sheet 4 — 费用明细
    # ===================================================================
    ws2 = wb.create_sheet("费用明细")
    for col_idx, h in enumerate(S2_HEADERS, 1):
        ws2.cell(row=1, column=col_idx, value=h)
    _style_header(ws2, 1, len(S2_HEADERS))

    row = 2
    for a in accruals:
        for fr in _extract_fee_rows(a, types_map):
            vals = [
                fr["accrual_id"], fr["date"], fr["posting_number"],
                fr["sku"], fr["type_id"], fr["type_name"],
                fr["amount"], fr["currency"], fr["source"],
            ]
            for col_idx, v in enumerate(vals, 1):
                _write_data_cell(ws2, row, col_idx, v, is_amount=(col_idx == 7))
            row += 1

    _auto_width(ws2, {
        1: 14, 2: 12, 3: 22, 4: 12, 5: 12, 6: 28, 7: 16, 8: 8, 9: 16,
    })
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions

    # ===================================================================
    # Sheet 5 — 商品佣金
    # ===================================================================
    ws3 = wb.create_sheet("商品佣金")
    for col_idx, h in enumerate(S3_HEADERS, 1):
        ws3.cell(row=1, column=col_idx, value=h)
    _style_header(ws3, 1, len(S3_HEADERS))

    row = 2
    for a in accruals:
        for pr in _extract_product_rows(a):
            vals = [
                pr["accrual_id"], pr["date"], pr["posting_number"],
                pr["sku"],
                pr["sale_amount"], pr["seller_price"],
                pr["sale_commission"], pr["commission"],
                pr["commission_ratio"], pr["bonus"],
                pr["coinvestment"], pr["delivery_total"],
                pr["currency"],
            ]
            for col_idx, v in enumerate(vals, 1):
                is_amt = col_idx in (5, 6, 7, 8, 10, 11, 12)
                _write_data_cell(ws3, row, col_idx, v, is_amount=is_amt)
            row += 1

    _auto_width(ws3, {
        1: 14, 2: 12, 3: 22, 4: 12, 5: 16, 6: 16, 7: 16,
        8: 14, 9: 10, 10: 14, 11: 16, 12: 14, 13: 8,
    })
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = ws3.dimensions

    # ===================================================================
    # Sheet 6 — 费用类型
    # ===================================================================
    ws4 = wb.create_sheet("费用类型")
    for col_idx, h in enumerate(S4_HEADERS, 1):
        ws4.cell(row=1, column=col_idx, value=h)
    _style_header(ws4, 1, len(S4_HEADERS))

    row = 2
    for tid in sorted(types_map.keys()):
        t = types_map[tid]
        _write_data_cell(ws4, row, 1, tid)
        _write_data_cell(ws4, row, 2, t["name"])
        _write_data_cell(ws4, row, 3, t["description"])
        row += 1

    _auto_width(ws4, {1: 10, 2: 35, 3: 60})
    ws4.freeze_panes = "A2"

    # ---- Save ----
    wb.save(filepath)
    print(f"Saved Excel report to {filepath}", file=sys.stderr)


# ===========================================================================
# Terminal summary
# ===========================================================================


def summarize(accruals: list[dict], types_map: dict[int, dict]) -> None:
    if not accruals:
        print("No accruals found.")
        return

    total_amount = 0.0
    cat_counts: dict[str, int] = {}
    for a in accruals:
        total_amount += _amt(a.get("total_amount"))
        cat = a.get("accrued_category", "UNKNOWN")
        cat_counts[cat] = cat_counts.get(cat, 0) + 1

    print(f"\n{'='*55}")
    print(f"  应计记录数: {len(accruals)}")
    print(f"  总金额:     {total_amount:>12,.2f}")
    print(f"\n  按类别统计:")
    for cat, cnt in sorted(cat_counts.items(), key=lambda x: -x[1]):
        print(f"    {cat:30s} {cnt:>6d}")
    print(f"  类型定义数: {len(types_map)}")
    print(f"{'='*55}")


# ===========================================================================
# Terminal summary — orders grouped by base posting number
# ===========================================================================


def print_order_summary(accruals: list[dict], types_map: dict[int, dict]) -> None:
    """Print a focused per-order summary to the terminal, most recent first.

    Same grouping logic as the "按订单号排列" Excel sheet:
      - Strips the trailing ``-N`` suffix from posting numbers.
      - Aggregates 收入 / 收单 / 销售佣金 / 配送佣金 / 国际配送 per base order.
    """
    import re

    def _base_pn(pn: str) -> str:
        return re.sub(r'-\d{1,2}$', '', str(pn))

    # 1. Build all focused rows (same as Excel sheet)
    all_focused: list[dict] = []
    for a in accruals:
        for fr in _build_focused_rows(a, types_map):
            all_focused.append(fr)

    if not all_focused:
        print("No data to summarize.")
        return

    # 2. Group by base posting number
    FIELDS = ["收入", "收单", "销售佣金", "配送佣金", "国际配送"]
    order_groups: dict[str, dict] = {}
    for fr in all_focused:
        pn = str(fr["posting_number"])
        base = _base_pn(pn)
        if base not in order_groups:
            order_groups[base] = {
                "收入": 0.0, "收单": 0.0, "销售佣金": 0.0,
                "配送佣金": 0.0, "国际配送": 0.0,
                "dates": set(), "skus": set(), "pns": set(),
            }
        g = order_groups[base]
        for field in FIELDS:
            g[field] += fr.get(field, 0.0) or 0.0
        if fr.get("date"):
            g["dates"].add(str(fr["date"]))
        sku = str(fr.get("sku", "-"))
        if sku and sku != "-":
            g["skus"].add(sku)
        g["pns"].add(pn)

    # 3. Sort by latest date descending (most recent first)
    def _sort_key(item: tuple[str, dict]) -> str:
        dates = item[1]["dates"]
        return max(dates) if dates else ""

    sorted_orders = sorted(order_groups.items(), key=_sort_key, reverse=True)

    # 4. Print table
    # Column widths
    W_PN = 24       # base posting number
    W_DATE = 12     # date
    W_SKU = 20      # SKUs
    W_AMT = 12      # amount columns
    HEADERS = ["基础单号", "日期", "SKU", "收入", "收单", "销售佣金", "配送佣金", "国际配送"]

    sep = "─" * (W_PN + W_DATE + W_SKU + W_AMT * 5 + 7 * 3 + 2)
    print(f"\n{sep}")
    print(
        f"  {HEADERS[0]:{W_PN}} │ {HEADERS[1]:{W_DATE}} │ {HEADERS[2]:{W_SKU}} │"
        f" {HEADERS[3]:>{W_AMT}} │ {HEADERS[4]:>{W_AMT}} │ {HEADERS[5]:>{W_AMT}} │"
        f" {HEADERS[6]:>{W_AMT}} │ {HEADERS[7]:>{W_AMT}}"
    )
    print(sep)

    total_row = {"收入": 0.0, "收单": 0.0, "销售佣金": 0.0, "配送佣金": 0.0, "国际配送": 0.0}

    for base, g in sorted_orders:
        dates = g["dates"]
        if len(dates) == 1:
            date_str = next(iter(dates))
        else:
            sd = sorted(dates)
            date_str = f"{sd[0]}~{sd[-1]}"
        sku_str = ", ".join(sorted(g["skus"])) if g["skus"] else "-"
        # Truncate long SKU strings
        if len(sku_str) > W_SKU:
            sku_str = sku_str[: W_SKU - 2] + "…"
        # Truncate long posting numbers
        pn_str = base if len(base) <= W_PN else base[: W_PN - 2] + "…"
        amts = [g[f] for f in FIELDS]

        print(
            f"  {pn_str:{W_PN}} │ {date_str:{W_DATE}} │ {sku_str:{W_SKU}} │"
            f" {amts[0]:{W_AMT},.2f} │ {amts[1]:{W_AMT},.2f} │ {amts[2]:{W_AMT},.2f} │"
            f" {amts[3]:{W_AMT},.2f} │ {amts[4]:{W_AMT},.2f}"
        )
        for i, f in enumerate(FIELDS):
            total_row[f] += g[f]

    print(sep)
    print(
        f"  {'合计':{W_PN}} │ {'':{W_DATE}} │ {'':{W_SKU}} │"
        f" {total_row['收入']:{W_AMT},.2f} │ {total_row['收单']:{W_AMT},.2f} │"
        f" {total_row['销售佣金']:{W_AMT},.2f} │ {total_row['配送佣金']:{W_AMT},.2f} │"
        f" {total_row['国际配送']:{W_AMT},.2f}"
    )
    print(sep)
    print(f"  {len(sorted_orders)} orders, {len(all_focused)} line items")


# ===========================================================================
# CLI
# ===========================================================================


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Fetch Ozon finance accruals via /v1/finance/accrual/by-day (LTS)"
    )
    parser.add_argument("--from", dest="date_from", help="Start date (YYYY-MM-DD). Default: 7 days ago.")
    parser.add_argument("--to", dest="date_to", help="End date (YYYY-MM-DD). Default: today.")
    parser.add_argument("--days", type=int, default=7, help="Days back from today. Default: 7.")
    parser.add_argument("--output", "-o", default="report.xlsx", help="Excel output path. Default: report.xlsx.")
    parser.add_argument("--json", help="Also dump raw JSON to this file.")
    parser.add_argument("--summary", "-s", action="store_true", help="Print per-order summary table to terminal.")
    parser.add_argument("--max-pages", type=int, default=100, help="Max cursor pages per day. Default: 100.")

    args = parser.parse_args()

    # --- Date range ---
    today = datetime.now(timezone.utc).date()
    date_to: str = args.date_to or today.isoformat()
    date_from: str = args.date_from or (today - timedelta(days=args.days)).isoformat()

    # --- Auth ---
    client_id, api_key = get_client()

    print(f"Date range : {date_from} → {date_to}", file=sys.stderr)
    print(f"Endpoints  : {TYPES_ENDPOINT}  +  {BY_DAY_ENDPOINT}  (v1 LTS)", file=sys.stderr)

    # --- 1. Fetch accrual types ---
    print("Fetching accrual types …", file=sys.stderr)
    types_map = fetch_accrual_types(client_id, api_key)
    print(f"  → {len(types_map)} types loaded", file=sys.stderr)
    # DEBUG: print all type names so we can verify keyword matching
    print("  Accrual types found:", file=sys.stderr)
    for tid in sorted(types_map.keys()):
        t = types_map[tid]
        print(f"    id={tid:>4}  name=\"{t['name']}\"  desc=\"{t['description']}\"", file=sys.stderr)

    # --- 2. Fetch accruals day by day ---
    accruals = fetch_all_accruals(client_id, api_key, date_from, date_to)

    # --- DEBUG: show all fees present in the data ---
    fee_summary: dict[str, tuple[int, float, str]] = {}  # type_name -> (count, total, source)
    for a in accruals:
        item_fees = a.get("item_fees") or {}
        for sg in (item_fees.get("fees") or []):
            for fee in (sg.get("fees") or []):
                tid = fee.get("type_id", 0)
                amt = _amt(fee.get("accrued"))
                tname = types_map.get(tid, {}).get("name", f"UNKNOWN_{tid}")
                prev = fee_summary.get(tname, (0, 0.0, "item_fee"))
                fee_summary[tname] = (prev[0] + 1, prev[1] + amt, "item_fee")
        nif = a.get("non_item_fee") or {}
        if nif:
            tid = nif.get("type_id", 0)
            amt = _amt(nif.get("accrued"))
            tname = types_map.get(tid, {}).get("name", f"UNKNOWN_{tid}")
            prev = fee_summary.get(tname, (0, 0.0, "non_item_fee"))
            fee_summary[tname] = (prev[0] + 1, prev[1] + amt, "non_item_fee")
        posting = a.get("posting") or {}
        for prod in (posting.get("products") or []):
            for svc in ((prod.get("delivery") or {}).get("services") or []):
                tid = svc.get("type_id", 0)
                amt = _amt(svc.get("accrued"))
                tname = types_map.get(tid, {}).get("name", f"UNKNOWN_{tid}")
                prev = fee_summary.get(tname, (0, 0.0, "delivery_svc"))
                fee_summary[tname] = (prev[0] + 1, prev[1] + amt, "delivery_svc")
    if fee_summary:
        print("  Fees found in data:", file=sys.stderr)
        for tname, (cnt, total, src) in sorted(fee_summary.items(), key=lambda x: -abs(x[1][1])):
            print(f"    {tname:35s} ×{cnt:>3}  total={total:>12.2f}  [{src}]", file=sys.stderr)
    else:
        print("  (no fees found in any accrual)", file=sys.stderr)

    # --- 3. Always Excel ---
    write_excel(accruals, types_map, args.output)

    # --- 4. Optional terminal summary ---
    if args.summary:
        print_order_summary(accruals, types_map)

    # --- Optional JSON ---
    if args.json:
        with open(args.json, "w", encoding="utf-8") as f:
            json.dump({"accruals": accruals, "types": types_map}, f, ensure_ascii=False, indent=2, default=str)
        print(f"Raw JSON saved to {args.json}", file=sys.stderr)

    # --- Summary ---
    summarize(accruals, types_map)


if __name__ == "__main__":
    main()
